"""
The module contains the main class and functionality 
for building various reports (academic performance).
"""
from asyncinit import asyncinit
import json
import os
from datetime import datetime
from enum import IntEnum
from pathlib import Path
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
from database.main_db import common_crud
from model.pydantic.home_work import DisciplineHomeWorks


class ReportFieldEnum(IntEnum):
    """The report column number in the generated file"""
    STUDENT_NAME = 1
    POINTS = 2
    LAB_COMPLETED = 3
    DEADLINES_FAILS = 4
    TASKS_COMPLETED = 5
    TASK_RATIO = 6
    NEXT = 7

@asyncinit
class BaseReportBuilder:
    """Base class that lays the framework of the report"""
    RED_FILL = PatternFill(
        start_color='FF0000',
        end_color='FF0000',
        fill_type='solid')
    GREEN_FILL = PatternFill(
        start_color='006633',
        end_color='006633',
        fill_type='solid')

    async def __init__(self,
                 group_id: int,
                 discipline_id: int,
                 prefix_file: str,
                 extension: str='xlsx'):
        """
        :param group_id: group id
        :param discipline_id: discipline id
        :param prefix_file: prefix of the name of the generated report
        :param extension: report file extension
        """
        self.group_id = group_id
        self.discipline_id = discipline_id
        self.wb = Workbook()
        # ====================================================================
        group = await common_crud.get_group(group_id)
        discipline = await common_crud.get_discipline(discipline_id)

        self.group_name = group.group_name
        self.discipline_name = discipline.short_name
        self.tasks_in_discipline = discipline.max_tasks

        path = Path(Path.cwd().joinpath(os.getenv("TEMP_REPORT_DIR")))
        self.__file_path = Path(
            path.joinpath(f'{discipline.short_name}_{prefix_file}_{group.group_name}_.{extension}')
        )
        return self

    async def build_report(self) -> None:
        """
        start creating and filling a basic report
        
        :return: None
        """
        worksheet = self.wb.active
        worksheet.title = self.discipline_name

        students = await common_crud.get_students_from_group(self.group_id)
        row = 1
        for student in students:
            assigned_discipline = await common_crud.get_disciplines_assigned_to_student(
                student.id,
                self.discipline_id)
            home_works = DisciplineHomeWorks(
                **json.loads(assigned_discipline.home_work)
                ).home_works
            if row == 1:
                worksheet.cell(row=row,
                               column=ReportFieldEnum.STUDENT_NAME
                               ).value = "ФИО студента"
                worksheet.cell(row=row,
                               column=ReportFieldEnum.POINTS
                               ).value = "Баллы (макс. 100)"
                worksheet.cell(row=row,
                               column=ReportFieldEnum.LAB_COMPLETED
                               ).value = "Полностью выполненых лаб"
                worksheet.cell(row=row,
                               column=ReportFieldEnum.DEADLINES_FAILS
                               ).value = "Кол-во сорванных сроков сдачи"
                worksheet.cell(row=row,
                               column=ReportFieldEnum.TASKS_COMPLETED
                               ).value = "Кол-во выполненных задач"
                worksheet.cell(row=row,
                               column=ReportFieldEnum.TASK_RATIO
                               ).value = "Объём выполненных задач"
                row += 1
            if row > 1:
                worksheet.cell(row=row,
                               column=ReportFieldEnum.STUDENT_NAME
                               ).value = student.full_name
                worksheet.cell(row=row,
                               column=ReportFieldEnum.POINTS
                               ).value = assigned_discipline.point

                deadlines_fails = 0
                lab_completed = 0
                for work in home_works:
                    if work.is_done:
                        lab_completed += 1
                    if datetime.now().date() > work.deadline:
                        if work.end_time is not None:
                            if work.end_time.date() > work.deadline:
                                deadlines_fails += 1
                        else:
                            deadlines_fails += 1
                worksheet.cell(row=row,
                               column=ReportFieldEnum.LAB_COMPLETED
                               ).value = lab_completed
                worksheet.cell(row=row,
                               column=ReportFieldEnum.DEADLINES_FAILS
                               ).value = deadlines_fails

                task_completed = 0
                for _, work in enumerate(home_works):
                    for _, task in enumerate(work.tasks):
                        if task.is_done:
                            task_completed += 1

                worksheet.cell(row=row,
                               column=ReportFieldEnum.TASKS_COMPLETED
                               ).value = task_completed
                worksheet.cell(
                    row=row, column=ReportFieldEnum.TASK_RATIO,
                ).value = round(task_completed / self.tasks_in_discipline, 3)

            row += 1

    def save_report(self) -> None:
        """
        Save the report
        
        :return: None
        """
        self.wb.save(self.get_path_to_report())

    def get_path_to_report(self) -> str:
        """
        Return the path to the generated report.

        :return: path to the generated report
        """
        return str(self.__file_path)
