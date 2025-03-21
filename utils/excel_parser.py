"""
This module contains everything you need to read and process data
about teachers and students from an Excel file.
"""
from enum import Enum
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from model.pydantic.db_start_data import StudentRaw, TeacherRaw


class ExcelDataParserError(Exception):
    """Any exceptions in the work of the Excel parser"""
    pass


class ParserType(Enum):
    """
    Data processing options
    
    :param ALL: 0 (Student and Teacher data)
    :param TEACHER: 1 (Teacher data)
    :param STUDENT: 2 (Student data)
    """
    ALL = 0
    TEACHER = 1
    STUDENT = 2


class ExcelDataParser:
    """
    For direct reading of data from an Excel file, data parsing,
    and saving for subsequent processing
    """
    def __init__(
            self, file_path: str, parse_type: ParserType = ParserType.ALL):
        """
        The constructor accepts a path to an Excel file containing information
        about teachers and students, as well as an optional parameter that
        specifies how this file should be processed.

        :param file_path: path to the excel file containing information
            about teachers and students
        :param parse_type: type of parsing used:
            for everything, teachers or students
        """
        # {discipline: {group_id: [students]}}
        self.__student: dict[str, dict[str, list[StudentRaw]]] = {}
        # {discipline: {group_id: [teachers]}}
        self.__teacher: dict[str, dict[str, list[TeacherRaw]]] = {}
        self.__parse_type = parse_type
        self.__load_data(file_path, parse_type)

    @property
    def students(self) -> dict[str, dict[str, list[StudentRaw]]]:
        """
        Get the dictionary of disciplines containing study groups and 
            a list of students studying in these groups
        
        :param None:

        :return dict[str, dict[str, list[StudentRaw]]]:
            {discipline: {group_id: [students]}}
        """
        if self.__parse_type == ParserType.TEACHER:
            raise ExcelDataParserError(
                "Students data don`t with this ParseType")
        return self.__student

    @property
    def teachers(self) -> dict[str, dict[str, list[TeacherRaw]]]:
        """
        Get a dictionary of disciplines containing study groups and
            a list of teachers teaching in these groups

        :param None:

        :return dict[str, dict[str, list[TeacherRaw]]]:
            {discipline: {group_id: [teachers]}}
        """
        if self.__parse_type == ParserType.STUDENT:
            raise ExcelDataParserError(
                "Teachers data don`t with this ParseType")
        return self.__teacher

    def __load_data(self, file_path: str, parse_type: ParserType) -> None:
        """
        Reads and writes information from a file,
            according to the selected processing type

        :param file_path: path to the excel file
        :param parse_type: the selected parsing type,
            eg ParserType.ALL, ParserType.TEACHER or ParserType.STUDENT

        :return None:
        """
        wb: Workbook = openpyxl.load_workbook(file_path)

        match parse_type:
            case ParserType.ALL:
                index = wb.sheetnames.index("teachers")
                wb.active = index
                self.__teachers_parser(wb.active)

                index = wb.sheetnames.index("students")
                wb.active = index
                self.__students_parser(wb.active)

            case ParserType.TEACHER:
                index = wb.sheetnames.index("teachers")
                wb.active = index
                self.__teachers_parser(wb.active)

            case ParserType.STUDENT:
                index = wb.sheetnames.index("students")
                wb.active = index
                self.__students_parser(wb.active)

            case _:
                raise ExcelDataParserError("ParserType not found")

    def __teachers_parser(self, worksheet: Worksheet) -> None:
        """
        Process information from teachers from an Excel file
        
        :param worksheet: the excel file worksheet that contains information
            about a teacher, their discipline and group

        :return None:
        """
        teacher_name = ""
        row = 2

        while teacher_name is not None:
            teacher_name = worksheet.cell(row=row, column=1).value
            telegram_id = worksheet.cell(row=row, column=2).value
            discipline = worksheet.cell(row=row, column=3).value
            is_admin = bool(worksheet.cell(row=row, column=4).value)
            group = worksheet.cell(row=row, column=5).value

            if teacher_name is None:
                break

            if discipline not in self.__teacher:
                self.__teacher[discipline] = {}

            if group not in self.__teacher[discipline]:
                self.__teacher[discipline][group] = []

            self.__teacher[discipline][group].append(
                TeacherRaw(
                    full_name=teacher_name,
                    telegram_id=telegram_id,
                    is_admin=is_admin
                )
            )

            row += 1

    def __students_parser(self, worksheet: Worksheet) -> None:
        """
        Process information from students from an Excel file

        :param worksheet: the excel file worksheet that contains information
            about a student, their discipline and group

        :return None:
        """
        student_name = ""
        row = 2

        while student_name is not None:
            student_name = worksheet.cell(row=row, column=1).value
            group = worksheet.cell(row=row, column=2).value
            discipline = worksheet.cell(row=row, column=3).value

            if student_name is None:
                break

            if discipline not in self.__student:
                self.__student[discipline] = {}

            if group not in self.__student[discipline]:
                self.__student[discipline][group] = []

            self.__student[discipline][group].append(
                StudentRaw(
                    full_name=student_name
                )
            )

            row += 1
